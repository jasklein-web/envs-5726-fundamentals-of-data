# Import necessary libraries for file paths, typing, CSV handling, date manipulation, numerical operations, plotting, Excel reading, and curve fitting
from pathlib import Path
from typing import List, Tuple, Dict
import csv
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.optimize import curve_fit

# TASK 1

# Define the path to the Excel file containing pipe material training data
pipe_material_excel_path = ('/Users/jasonklein/Downloads/Pipe_Material_Training_Data.xlsx')

# Load the Excel workbook into memory
workbook = load_workbook(pipe_material_excel_path)

# Access the specific sheet named "Survival Probabilities"
sheet = workbook["Survival Probabilities"]

# Define the Weibull Cumulative Distribution Function (CDF)
# This function models the probability of failure by a given age
def cumulative_density_function(x, c, b, a):
    # x = age, a = shape parameter (controls curve steepness),
    # b = scale parameter (controls lifespan length),
    # c = initial survival multiplier (vertical scaling)
    return 1 - c * np.exp(-((x / b) ** a))  # Returns failure probability by age x

# Define a function to fit and plot the Weibull CDF for a specific pipe material
def plot_cdf_curve_fit(table, material_type, line_color):
    X_train = []  # List to store pipe ages (Life Expectancy)
    Y_train = []  # List to store failure probabilities (CDF values)

    # Iterate through each row in the Excel sheet
    for row in table.iter_rows(values_only=True):
        # Unpack row into material, life expectancy, and survival percentage
        mat, life, surv = row
        # Skip the header row
        if mat == "Material Type":
            continue

        # Only process rows matching the specified material type
        if mat == material_type:
            # Skip invalid rows where life expectancy is non-positive
            if float(life) <= 0:
                continue

            # Convert survival percentage to failure probability (CDF)
            cdf_value = 1 - float(surv) / 100

            # Add to training lists
            X_train.append(float(life))
            Y_train.append(cdf_value)

    # Convert lists to NumPy arrays for numerical operations
    X_train = np.array(X_train)
    Y_train = np.array(Y_train)

    # Handle case where all Y values are identical (e.g., Copper data)
    # Add small noise to allow curve fitting
    if np.all(Y_train == Y_train[0]):
        Y_train = Y_train + 0.001 * np.arange(len(Y_train))

    # Fit the Weibull model to the training data
    coeffs, _ = curve_fit(
        cumulative_density_function,  # The model function
        X_train,  # Age data
        Y_train,  # Failure probability data
        p0=[1, 50, 2],  # Initial guess for c, b, a
        maxfev=20000  # Maximum function evaluations
    )

    # Unpack fitted coefficients
    c, b, a = coeffs
    # Print fitted parameters for debugging/verification
    print(f"{material_type}: c={c:.4f}, b={b:.4f}, a={a:.4f}")

    # Generate smooth curve for plotting
    x_vals = np.linspace(1, max(X_train) * 1.2, 300)  # Age range from 1 to slightly beyond max training age
    y_vals = cumulative_density_function(x_vals, c, b, a)  # Compute failure probabilities

    # Convert failure probability back to survival percentage for visualization
    survival_vals = (1 - y_vals) * 100

    # Plot the survival curve
    plt.plot(x_vals, survival_vals, color=line_color, label=material_type)

# Fit and plot survival curves for each material type
plot_cdf_curve_fit(sheet, "Cast Iron", "red")
plot_cdf_curve_fit(sheet, "Ductile Iron", "blue")
plot_cdf_curve_fit(sheet, "Galvanized Iron", "green")
plot_cdf_curve_fit(sheet, "Copper", "brown")

# Add chart elements
plt.legend()
plt.xlabel("Life Expectancy (y)")
plt.ylabel("Survival Probability (%)")
plt.title("Pipe Material Survival CDF")
plt.grid(True)
plt.show()

# TASK 2
# Import additional libraries for data structures and math operations
from collections import namedtuple
import math

# Define a function to compute survival probability using Weibull formula
def weibull_survival(age, c, b, a):
    return c * math.exp(-((age / b) ** a)) * 100  # Returns survival probability as a percentage

# Dictionary of fitted Weibull coefficients from Task 1 for each material
coefficients = {
    "Cast Iron": (1.0236, 91.2607, 1.6987),
    "Ductile Iron": (1.0434, 66.0815, 1.6726),
    "Galvanized Iron": (1.2914, 25.9335, 1.4309),
    "Copper": (1.0903, 44.1999, 1.6471)
}

# Define a namedtuple to represent a water main record (immutable, readable)
WaterMain = namedtuple(
    "WaterMain",
    ["MainType", "Diameter", "InstallDate", "Material", "Age", "Survival_Probability"]
)

# List to store all water main records
water_mains_table = []

# Path to the CSV file containing water main data
csv_path = '/Users/jasonklein/Downloads/Water_Mains.csv'

# Open and read the CSV file
with open(csv_path, newline='', encoding="windows-1252") as csvfile:
    reader = csv.DictReader(csvfile)  # Read rows as dictionaries
    current_year = 2025  # Reference year for age calculation

    for row in reader:
        # Fix encoding issue with column name (BOM character)
        if "ï»¿MainType" in row:
            row["MainType"] = row.pop("ï»¿MainType")

        # Parse the installation date string into a datetime object
        install_date = datetime.strptime(row["InstallDate"], "%m/%d/%Y %H:%M")
        material = row["Material"]

        # Calculate age in years
        age = current_year - install_date.year

        # Calculate survival probability using material-specific coefficients
        if material in coefficients:
            c, b, a = coefficients[material]
            survival = weibull_survival(age, c, b, a)
            # Cap survival probability at 100%
            if survival > 100:
                survival = 100
        else:
            survival = None  # If material not in dictionary, set to None

        # Create a WaterMain namedtuple instance for this pipe
        wm = WaterMain(
            MainType=row["MainType"],
            Diameter=row["Diameter"],
            InstallDate=install_date,
            Material=material,
            Age=age,
            Survival_Probability=round(survival, 3) if survival is not None else None
        )

        # Add to the table
        water_mains_table.append(wm)

# Print first 5 rows to verify data
for row in water_mains_table[:5]:
    print(row)

# TASK 3

# Extract survival probabilities from the water mains table (excluding None values)
survival_values = [
    row.Survival_Probability
    for row in water_mains_table
    if row.Survival_Probability is not None
]

# Define threshold for "low survival" (below 70% is considered vulnerable)
threshold = 70

# Filter list for pipes with survival probability below threshold
low_survival = [v for v in survival_values if v < threshold]

# Calculate percentage of pipes below threshold
percent_low = (len(low_survival) / len(survival_values)) * 100

# Create a histogram of survival probabilities
plt.figure(figsize=(8, 5))
plt.hist(survival_values, bins=15, color='skyblue', edgecolor='black')
# Add a vertical line at the threshold
plt.axvline(threshold, color='red', linestyle='--', label=f'{threshold}% threshold')

# Add chart labels and title
plt.title('Distribution of Pipe Survival Probabilities – Mercator Water')
plt.xlabel('Survival Probability (%)')
plt.ylabel('Number of Pipes')
plt.legend()
plt.grid(axis='y', alpha=0.3)
plt.show()

# Print summary statistics
print(f"Pipes below {threshold}% survival: {len(low_survival)} ({percent_low:.1f}% of total)")

"""
Based on the histogram of pipe survival probabilities for Mercator Water, the utility does not appear to be vulnerable to 
overwhelming pipe failure. The distribution is heavily skewed toward high survival rates, with the majority of
pipes exhibiting survival probabilities above 70%—many clustering near 80–100%. Only a small fraction of pipes fall below 
the 70% threshold, indicating that the overall system remains in good condition. This suggests that the risk of widespread
infrastructure failure is low, and Mercator Water can likely manage necessary repairs or replacements within its current 
budget without needing to sell to a private utility.
"""